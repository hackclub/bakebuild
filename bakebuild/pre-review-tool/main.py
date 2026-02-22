import os
import re
import requests
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed, TimeoutError
import multiprocessing
from collections import defaultdict
from threading import Lock

load_dotenv()

AIRTABLE_API_KEY = os.getenv("AIRTABLE_API_KEY")
BASE_ID = os.getenv("BASE_ID")
TABLE_ID = os.getenv("TABLE_ID")
SUBMISSIONS_VIEW = os.getenv("SUBMISSIONS_VIEW")
DUPLICATE_VIEW = os.getenv("DUPLICATE_VIEW")
WORKSHOP_TABLE_ID = os.getenv("WORKSHOP_TABLE_ID")
WORKSHOP_NAME_FIELD = "Name"

STEP_FIELD = "Project file (STEP)"
FIRSTNAME_FIELD = "First Name"
LASTNAME_FIELD = "Last Name"
ADDRESS_FIELD = "Address (Line 1)"
WORKSHOP_FIELD = "Workshop"

OUTPUT_DIR = Path("downloaded_step_files")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

LIMIT_MM = 100.0
HEADERS = {"Authorization": f"Bearer {AIRTABLE_API_KEY}"}

STEP_POOL = ProcessPoolExecutor(max_workers=max(2, multiprocessing.cpu_count() - 1))


def fetch_airtable_records_for_view(view_id):
    url = f"https://api.airtable.com/v0/{BASE_ID}/{TABLE_ID}"
    params = {"view": view_id, "pageSize": 100}
    records = []
    while True:
        resp = requests.get(url, headers=HEADERS, params=params)
        resp.raise_for_status()
        data = resp.json()
        records.extend(data.get("records", []))
        offset = data.get("offset")
        if not offset:
            break
        params["offset"] = offset
    return records


def fetch_airtable_table(table_id):
    url = f"https://api.airtable.com/v0/{BASE_ID}/{table_id}"
    params = {"pageSize": 100}
    records = []
    while True:
        resp = requests.get(url, headers=HEADERS, params=params)
        resp.raise_for_status()
        data = resp.json()
        records.extend(data.get("records", []))
        offset = data.get("offset")
        if not offset:
            break
        params["offset"] = offset
    return records


def _parse_step_file(path: str):
    from build123d import import_step
    shape = import_step(path)
    bb = shape.bounding_box()
    return (float(bb.size.X), float(bb.size.Y), float(bb.size.Z))


def is_step_file(filename: str) -> bool:
    if not filename:
        return False
    return filename.lower().endswith((".step", ".stp"))


def safe_step_dimensions(filepath, timeout=30):
    future = STEP_POOL.submit(_parse_step_file, str(filepath))
    try:
        return future.result(timeout=timeout)
    except TimeoutError:
        future.cancel()
        raise RuntimeError("STEP parsing timed out")
    except Exception as e:
        raise RuntimeError(f"STEP parsing failed: {e}")


def download_worker(
    rec,
    workshop_map,
    dup_names,
    dup_addresses,
    submitted_count_by_person,
    too_many_people,
    lock
):
    fields = rec.get("fields", {})
    firstname = fields.get(FIRSTNAME_FIELD)
    lastname = fields.get(LASTNAME_FIELD)
    address = fields.get(ADDRESS_FIELD, "")
    workshop_ids = fields.get(WORKSHOP_FIELD, [])
    attachments = fields.get(STEP_FIELD)
    results = []

    if not firstname or not lastname or not attachments:
        return results

    person_key = (
        firstname.strip().lower(),
        lastname.strip().lower(),
        address.strip().lower()
    )

    workshop_names = [workshop_map.get(wid, "") for wid in workshop_ids]
    workshop_text = ", ".join([w for w in workshop_names if w])

    is_dup = (
        (firstname.strip().lower(), lastname.strip().lower()) in dup_names
        or address.strip().lower() in dup_addresses
    )

    fname_safe = re.sub(r"\s+", "-", firstname.strip())
    lname_safe = re.sub(r"\s+", "-", lastname.strip())

    local_idx = 0

    for att in attachments:
        with lock:
            if submitted_count_by_person[person_key] >= 3:
                too_many_people.add(person_key)
                break
            submitted_count_by_person[person_key] += 1

        local_idx += 1
        idx = local_idx

        url = att.get("url")
        original_name = att.get("filename", "")

        if not url:
            continue

        valid_step = is_step_file(original_name)

        if valid_step:
            filename = f"{fname_safe}_{lname_safe}_cutter{idx}.step"
        else:
            ext = Path(original_name).suffix or ".invalid"
            filename = f"{fname_safe}_{lname_safe}_cutter{idx}{ext}"

        filepath = OUTPUT_DIR / filename

        try:
            resp = requests.get(url, timeout=30)
            resp.raise_for_status()
            with open(filepath, "wb") as fh:
                fh.write(resp.content)
        except Exception:
            continue

        results.append({
            "first": firstname,
            "last": lastname,
            "addr": address,
            "workshop": workshop_text,
            "filename": filename,
            "filepath": filepath,
            "dup": is_dup,
            "too_many": False,
            "invalid_file": not valid_step
        })

    if person_key in too_many_people:
        for r in results:
            r["too_many"] = True

    return results


def main():
    submissions = fetch_airtable_records_for_view(SUBMISSIONS_VIEW)
    duplicates_pool = fetch_airtable_records_for_view(DUPLICATE_VIEW)
    workshop_recs = fetch_airtable_table(WORKSHOP_TABLE_ID)

    workshop_map = {
        r["id"]: r.get("fields", {}).get(WORKSHOP_NAME_FIELD, "")
        for r in workshop_recs
    }

    dup_names = set()
    dup_addresses = set()
    for r in duplicates_pool:
        f = r.get("fields", {}).get(FIRSTNAME_FIELD, "") or ""
        l = r.get("fields", {}).get(LASTNAME_FIELD, "") or ""
        a = r.get("fields", {}).get(ADDRESS_FIELD, "") or ""

        if f and l:
            dup_names.add((f.strip().lower(), l.strip().lower()))
        if a:
            dup_addresses.add(a.strip().lower())

    submitted_count_by_person = defaultdict(int)
    too_many_people = set()
    lock = Lock()

    download_results = []
    with ThreadPoolExecutor(max_workers=8) as ex:
        futures = [
            ex.submit(
                download_worker,
                rec,
                workshop_map,
                dup_names,
                dup_addresses,
                submitted_count_by_person,
                too_many_people,
                lock
            )
            for rec in submissions
        ]

        for f in as_completed(futures):
            try:
                download_results.extend(f.result())
            except Exception:
                pass

    final_rows = []

    for item in download_results:
        filename = item["filename"]
        filepath = item["filepath"]

        if item.get("invalid_file"):
            dx = dy = dz = None
            fits = False
            fits_label = "INVALID"
        else:
            try:
                dx, dy, dz = safe_step_dimensions(filepath)
                fits = dx <= LIMIT_MM and dy <= LIMIT_MM and dz <= LIMIT_MM
                fits_label = "YES" if fits else "NO"
            except Exception:
                dx = dy = dz = None
                fits = False
                fits_label = "ERROR"

        final_rows.append({
            "First Name": item["first"],
            "Last Name": item["last"],
            "Address": item["addr"],
            "Workshop": item["workshop"],
            "Filename": filename,
            "Size X (mm)": f"{dx:.2f}" if dx is not None else "",
            "Size Y (mm)": f"{dy:.2f}" if dy is not None else "",
            "Size Z (mm)": f"{dz:.2f}" if dz is not None else "",
            "FitsFlag": fits,
            "Fits 100mm³?": fits_label,
            "DuplicateFlag": item["dup"],
            "Duplicate?": "YES" if item["dup"] else "NO",
            "TooManyFlag": item["too_many"],
            "More than 3 Cutters?": "YES" if item["too_many"] else "NO",
            "InvalidFileFlag": item.get("invalid_file", False),
            "Invalid File Type?": "YES" if item.get("invalid_file") else "NO"
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "Submissions"

    headers = [
        "First Name", "Last Name", "Address", "Workshop",
        "Filename", "Size X (mm)", "Size Y (mm)", "Size Z (mm)",
        "Fits 100mm³?", "Duplicate?", "More than 3 Cutters?",
        "Invalid File Type?"
    ]
    ws.append(headers)

    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ORANGE_FILL = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    PURPLE_FILL = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")

    row_index = 2
    for r in final_rows:
        ws.append([
            r["First Name"], r["Last Name"], r["Address"], r["Workshop"],
            r["Filename"], r["Size X (mm)"], r["Size Y (mm)"], r["Size Z (mm)"],
            r["Fits 100mm³?"], r["Duplicate?"], r["More than 3 Cutters?"],
            r["Invalid File Type?"]
        ])

        row = ws[row_index]

        if r["InvalidFileFlag"]:
            for cell in row:
                cell.fill = PURPLE_FILL
        elif r["TooManyFlag"]:
            for cell in row:
                cell.fill = ORANGE_FILL
        elif r["DuplicateFlag"]:
            for cell in row:
                cell.fill = RED_FILL
        elif r["Fits 100mm³?"] == "ERROR":
            for cell in row:
                cell.fill = ORANGE_FILL
        elif not r["FitsFlag"]:
            for cell in row:
                cell.fill = YELLOW_FILL

        row_index += 1

    wb.save("submission_results.xlsx")
    STEP_POOL.shutdown(cancel_futures=True)


if __name__ == "__main__":
    main()